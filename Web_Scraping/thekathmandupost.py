from bs4 import BeautifulSoup
from selenium import webdriver
from time import sleep
import os
import requests
from fake_useragent import UserAgent


driver=webdriver.Chrome('chromedriver')
main_url='https://kathmandupost.com'
driver.get(main_url)

#%%
sleep(10)
#go to business section
business_section=driver.find_element_by_link_text('Money')
business_section.click()
#%%
user_agent=UserAgent()
business_section_url='https://kathmandupost.com/money'
page=requests.get(business_section_url,headers={'user-agent':user_agent.chrome})

soup=BeautifulSoup(page.content,'lxml')
#%%
all_divs=soup.find('div',class_='block--morenews')
for h3 in all_divs.find_all('h3'):
    print(h3.text)
for a in all_divs.find_all('a'):
    print(a['href'])
#%%
titles=[h3.text for h3 in all_divs.find_all('h3')]
links=[main_url+a['href'] for a in all_divs.find_all('a')]
sliced_links=links[0:len(links):3]
#%%
"""from xlsxwriter import Workbook

#make workbook
workbook=Workbook('news-english.xlsx')

#add worksheet
worksheet=workbook.add_worksheet()
worksheet.write(0,0,"Titles")
worksheet.write(0,1,"Links")
for i in range(len(titles)):
    worksheet.write(i+1,0,titles[i])
    worksheet.write(i+1,1,sliced_links[i])

workbook.close()"""
#%%
from openpyxl import load_workbook

book=load_workbook('news-english.xlsx')
sheet=book.active

for i in range(len(titles)):
    sheet.append([titles[i],sliced_links[i]])
    
book.save('news-english.xlsx')
#%%
#handle duplicates
import pandas as pd
file_df = pd.read_excel("news-english.xlsx")
file_df_first_record = file_df.drop_duplicates(subset=["Titles", "Links"], keep="first")
file_df_first_record.to_excel("news-english.xlsx", index=False)

