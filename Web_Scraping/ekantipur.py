from bs4 import BeautifulSoup
from selenium import webdriver
from time import sleep
import os
import requests
from fake_useragent import UserAgent


driver=webdriver.Chrome('chromedriver')
main_url='https://ekantipur.com'
driver.get(main_url)
#%%
sleep(10)
#cancel popup dialog
cancel_button=driver.find_element_by_id('onesignal-slidedown-cancel-button')
cancel_button.click()
sleep(5)
#extend the sections
icon_hamburger=driver.find_element_by_xpath('//*[@id="header"]/div[1]/nav/label/i')
icon_hamburger.click()
sleep(5)
#go to business section
business_section=driver.find_element_by_link_text('अर्थ / वाणिज्य')
business_section.click()

#%%
user_agent=UserAgent()
business_section_url='https://ekantipur.com/business'
page=requests.get(business_section_url,headers={'user-agent':user_agent.chrome})

soup=BeautifulSoup(page.content,'lxml')
#%%
"""all_divs=soup.find_all('div',class_='col-xs-10 col-sm-10 col-md-10')

for div in all_divs:
    print(div)"""
all_articles=soup.find_all('article',class_='normal')
for article in all_articles:    
    print(article.a)
#%%   
titles=[article.a.string for article in all_articles]
for title in titles:
    print(title)
    
title_links=[article.a['href'] for article in all_articles]
for link in title_links:
    print(main_url+link)
#%%

all_titles=[article.a.string for article in all_articles]    
all_links=[main_url+article.a['href'] for article in all_articles]

#%%"""
#save into an excel file
"""from xlsxwriter import Workbook

#make workbook
workbook=Workbook('news.xlsx')

#add worksheet
worksheet=workbook.add_worksheet()
worksheet.write(0,0,"Titles")
worksheet.write(0,1,"Links")
for i in range(len(all_titles)):
    worksheet.write(i+1,0,all_titles[i])
    worksheet.write(i+1,1,all_links[i])

workbook.close()"""
#%%
from openpyxl import load_workbook

book=load_workbook('news.xlsx')
sheet=book.active

for i in range(len(all_titles)):
    sheet.append([all_titles[i],all_links[i]])
    
book.save('news.xlsx')
#%%
#handle duplicates
import pandas as pd
file_df = pd.read_excel("news.xlsx")
file_df_first_record = file_df.drop_duplicates(subset=["Titles", "Links"], keep="first")
file_df_first_record.to_excel("news.xlsx", index=False)




